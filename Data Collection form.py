from openpyxl import *
from tkinter import *

wb = load_workbook("data.xlsx")
sheet = wb.active

def focus0(event):
    name_field.focus_set()
def focus1(event):
    course_field.focus_set()
def focus2(event):
    sem_field.focus_set()
def focus3(event):
    form_no_field.focus_set()
def focus4(event):
    contact_no_field.focus_set()
def focus5(event):
    email_id_field.focus_set()
def focus6(event):
    address_field.focus_set()

def clear():
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)

def insert():
    if (name_field.get()==""and
    course_field.get()==""and
    sem_field.get()==""and
    form_no_field.get()==""and
    contact_no_field.get()==""and
    email_id_field.get()==""and
    address_field.get()==""):
        print("Empty Field")
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()
        wb.save("data.xlsx")
        name_field.focus_set()
        clear()

if __name__ == "__main__":
    root = Tk()
    root.title("DATA COLLECTION FORM")
    root.geometry("500x330+400+180")
    root.configure(background='black')
    
    heading = Label(root, text="Form", fg="white",bg="black")
    name = Label(root, text= "Name", fg="white",bg="black")
    course = Label(root,text= "Occupation", fg="white",bg="black")
    sem = Label(root,text= "Job title", fg="white",bg="black")
    form_no = Label(root, text= "ID no.", fg="white",bg="black")
    contact_no = Label(root, text= "Contact Number", fg="white",bg="black")
    email_id = Label(root, text= "Email-ID", fg="white",bg="black")
    address = Label(root, text= "Address", fg="white",bg="black")
     
    name.grid(row=1, column=0,padx=10,pady=10)
    course.grid(row=2, column=0,padx=10,pady=10)
    sem.grid(row=3, column=0,padx=10,pady=10)
    form_no.grid(row=4, column=0,padx=10,pady=10)
    contact_no.grid(row=5, column=0,padx=10,pady=10)
    email_id.grid(row=6, column=0,padx=10,pady=10)
    address.grid(row=7, column=0,padx=10,pady=10)
    name_field = Entry(root)
    course_field=Entry(root)
    sem_field=Entry(root)
    form_no_field=Entry(root)
    contact_no_field=Entry(root)
    email_id_field=Entry(root)
    address_field=Entry(root)
    name_field.bind("&lt;Return&gt;",focus0)
    course_field.bind("&lt;Return&gt;",focus1)
    sem_field.bind("&lt;Return&gt;", focus2)
    form_no_field.bind("&lt;Return&gt;", focus3)
    contact_no_field.bind("&lt;Return&gt;",focus4)
    email_id_field.bind("&lt;Return&gt;", focus5)
    address_field.bind("&lt;Return&gt;", focus6)        

    name_field.grid(row=1, column=1, ipadx="100",padx=10,pady=10)
    course_field.grid(row=2, column=1, ipadx="100",padx=10,pady=10)
    sem_field.grid(row=3, column=1, ipadx="100",padx=10,pady=10)
    form_no_field.grid(row=4, column=1, ipadx="100",padx=10,pady=10)
    contact_no_field.grid(row=5, column=1, ipadx="100",padx=10,pady=10)
    email_id_field.grid(row=6, column=1, ipadx="100",padx=10,pady=10)
    address_field.grid(row=7, column=1, ipadx="100",padx=10,pady=10)
    
    submit = Button(root, text="Submit", fg="white",bg="black", command=insert)
    submit.grid(row=8, column=1,padx=10,pady=10)
    root.mainloop()
